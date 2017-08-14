from flask import Flask, send_from_directory, after_this_request, render_template, request
import praw
from datetime import datetime
from openpyxl import Workbook, styles
from urllib.parse import urlparse
import os


app = Flask(__name__)

#   Sara's Reddit Scraper
#       * Because you can totally just build
#         a thing in a day that does what you want.

r = praw.Reddit(client_id=os.environ.get('PRAW_CLIENT_ID'),
                client_secret=os.environ.get('PRAW_CLIENT_SECRET'),
                user_agent='PRAW/Python 3.6.2')


# Helpers

def parse_rawurl(url):
    path = urlparse(url).path.split('/')
    id = path[4]
    return id


def format_spreadsheet(ws):

    # freeze top row
    ws.freeze_panes = ws['A2']

    # set top row to bold
    # we can't address the whole row at once, gotta do cells.
    # not going to bother with a content checker
    top_row = [ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1']]
    for r in top_row:
        r.font = styles.Font(bold=True)

    dims = {}
    for row in ws.rows:
        for cell in row:
            # set length of cells to length of content, with 100 char limit.
            if cell.value:
                if len(cell.value) < 100:
                    cell_length = len(cell.value) + 1
                else:
                    cell_length = 100
                dims[cell.column] = max((dims.get(cell.column, 0), cell_length))
                # Set alignment to wrap text, this preserves newlines.
                cell.alignment = styles.Alignment(wrap_text=True)
    for col, value in dims.items():
        ws.column_dimensions[col].width = value
    return ws


# Data Scraping / Normalization Interfaces

class ScrapedSubmission:

    # Submission object fields are fetched from API when they are addressed/called.
    def __init__(self, r, submission_id):

        self.post = r.submission(id=submission_id)

        self.title = self.post.title
        self.upvotes = str(self.post.ups)
        self.downvotes = str(self.post.downs)
        self.sub = self.post.subreddit.display_name
        self.selftext = self.post.selftext
        self.url = self.post.url

        self.created_at = datetime.fromtimestamp(
            int(self.post.created_utc)
        ).strftime('%Y-%m-%d %H:%M:%S UTC')

        self.post_url = 'https://www.reddit.com' + self.post.permalink

        if self.post.author:
            self.author = self.post.author.name
        else:
            self.author = '[deleted]'

        # We don't need to wait around, lets just get them now.
        self.comments = self._get_comments()

    def _get_comments(self):

        scraped = []

        comments = self.post.comments

        # expand MoreComments objects
        comments.replace_more()

        for comment in comments.list():
            scraped.append(ScrapedComment(comment))

        return scraped


class ScrapedComment:

    def __init__(self, commentobj):

        self.c = commentobj

        self.body = self.c.body
        self.upvotes = str(self.c.ups)

        self.created_at = datetime.fromtimestamp(
            int(self.c.created_utc)
        ).strftime('%Y-%m-%d %H:%M:%S UTC')

        self.permalink = "https://www.reddit.com" +  self.c.permalink(fast=True)

        if hasattr(self.c.author, 'name'):
            self.username = self.c.author.name
        else:
            self.username = '[deleted]'


def excel_writer(filename, scraped_submission):

    ss = scraped_submission

    # Sort comments by date
    comments = sorted(ss.comments, key=lambda comment: comment.created_at)

    # Intialize workbook in memory
    wb = Workbook()
    comment_data = wb.active
    comment_data.title = 'Comment Data'
    post_metadata = wb.create_sheet(title='Post Metadata')

    # Write metadata
    post_metadata.append(['Attribute', 'Data'])
    post_metadata.append(['Post Title', ss.title])
    post_metadata.append(['Post Author', ss.author])
    post_metadata.append(['Post Created At', ss.created_at])
    post_metadata.append(['Submitted URL', ss.url])
    post_metadata.append(['Post Self Text', ss.selftext])
    post_metadata.append(['Link to Post', ss.post_url])

    # Write comments data
    comment_data.append(['Date Posted', 'Reddit User', 'Upvotes', 'Link', 'Comment Body'])
    for c in comments:
        comment_data.append([c.created_at, c.username, c.upvotes, c.permalink, c.body])

    # set column width, text wrapping
    for ws in post_metadata, comment_data:
        format_spreadsheet(ws)

    # Write file
    wb.save(filename)


### views

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static/favicon'),
                               'favicon.ico', mimetype='image/vnd.microsoft.icon')

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/get-submission', methods=['POST'])
def scrape_submission_and_send():
    rawurl = request.form['rawurl']
    sid = parse_rawurl(rawurl)
    filename = '{}.xlsx'.format(sid)

    excel_writer('/tmp/' + filename, ScrapedSubmission(r, sid))

    @after_this_request
    def cleanup_temp(response):
        os.remove('/tmp/' + filename)
        return response

    return send_from_directory(directory='/tmp', filename=filename, as_attachment='True')


if __name__ == '__main__':
    if os.environ.get('PORT'):
        host = '0.0.0.0'
        debug = False
    else:
        host = '127.0.0.1'
        debug = True
    app.run(port=os.environ.get('PORT'), host=host, debug=debug)
